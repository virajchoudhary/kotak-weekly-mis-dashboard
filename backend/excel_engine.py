from __future__ import annotations

from copy import copy
import math
import os
from pathlib import Path
import re
import shutil
import subprocess
import tempfile
from zipfile import ZIP_DEFLATED, ZipFile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.views import Selection

from .constants import SUMMARY_METRICS
from .mapping import is_fintech_row


EXPECTED_SHEETS = [
    "Summary - Banks, ND & RIA ",
    "Summary - FINTECH",
    "SIP Pivot",
    "Brokerwise Data",
]

AMOUNT_FMT = "#,##0.00"
COUNT_FMT = "#,##0"
SHARE_FMT = "0.00%"
FORMULA_ERRORS = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?"}

def apply_safe_excel_view(ws, active_cell="A1"):
    """Make a worksheet open cleanly near the top-left with no frozen panes.

    View-only: resets the saved scroll position, active/selected cell, zoom and view
    mode and removes frozen panes.
    It never touches cell values, formulas, styles, merges, widths, or number formats.
    """
    view = ws.sheet_view
    view.view = "normal"            # not pageBreakPreview / pageLayout
    view.topLeftCell = None          # clear any saved scroll position -> opens at A1
    view.zoomScale = None            # 100%
    view.zoomScaleNormal = None
    view.tabSelected = False         # the active sheet is set by apply_safe_workbook_views
    view.selection = [Selection(activeCell=active_cell, sqref=active_cell)]
    ws.freeze_panes = None


def apply_safe_workbook_views(workbook) -> None:
    """Reset every generated sheet to a clean, header-visible view and open the
    first sheet first. Safe to call after the workbook is fully built."""
    for index, ws in enumerate(workbook.worksheets):
        apply_safe_excel_view(ws)
        ws.sheet_view.tabSelected = index == 0
    workbook.active = 0


def safe_ratio(numerator: float, denominator: float) -> float:
    return float(numerator) / float(denominator) if denominator else 0.0


_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def _safe_cell_text(value: object) -> str:
    """Neutralise spreadsheet formula/DDE injection in user-supplied text cells.

    openpyxl turns a string beginning with ``=`` into a live formula, so a
    malicious broker/ARN value (e.g. ``=cmd|'/c calc'!A1`` or ``=WEBSERVICE(...)``)
    could execute or exfiltrate when an analyst opens the downloaded workbook.
    Prefixing a literal apostrophe forces Excel to store it as plain text.
    """
    text = "" if value is None else str(value)
    if text[:1] in _FORMULA_TRIGGERS:
        return "'" + text
    return text


def build_summary_rows(
    frame: pd.DataFrame, scheme_master: list[dict], fintech: bool
) -> list[dict]:
    """Build the primary overall summary or the separate FINTECH breakout.

    The primary 45-scheme summary intentionally includes every brokerwise row,
    including FINTECH, so its grand totals tie directly to Brokerwise Data.
    """
    if frame.empty:
        selected = frame
    else:
        fintech_mask = frame.apply(is_fintech_row, axis=1)
        selected = frame[fintech_mask] if fintech else frame

    order_key = "display_order_fintech" if fintech else "display_order_banks_nd_ria"
    include_key = "include_in_fintech" if fintech else "include_in_banks_nd_ria"
    master_rows = sorted(
        [row for row in scheme_master if row[include_key]],
        key=lambda row: row[order_key],
    )
    result: list[dict] = []
    for master in master_rows:
        subset = selected[selected["asset_class"] == master["asset_class"]]
        record = {
            "asset_class": master["asset_class"],
            "sch_group": master["sch_group"],
        }
        for kotak, cams, ms in SUMMARY_METRICS:
            record[kotak] = float(subset[kotak].sum()) if not subset.empty else 0.0
            record[cams] = float(subset[cams].sum()) if not subset.empty else 0.0
            record[ms] = safe_ratio(record[kotak], record[cams])
        result.append(record)
    return result


def build_sip_pivot(frame: pd.DataFrame) -> list[dict]:
    if frame.empty:
        return []
    grouped = (
        frame.groupby(["arn_code", "broker_name"], dropna=False)[
            ["kotak_sip_count", "cams_sip_count"]
        ]
        .sum()
        .reset_index()
        .sort_values(["arn_code", "broker_name"])
    )
    return grouped.to_dict(orient="records")


def _sheet_by_trimmed_name(workbook, name: str):
    for worksheet in workbook.worksheets:
        if worksheet.title.strip() == name.strip():
            return worksheet
    raise ValueError(f"Template is missing sheet: {name}")


def _write_summary_sheet(worksheet, rows: list[dict]) -> None:
    first_row = 3
    expected_last = first_row + len(rows) - 1
    total_row = expected_last + 2
    if worksheet.max_row != total_row:
        raise ValueError(
            f"Template row count for {worksheet.title} is {worksheet.max_row}; expected {total_row}"
        )
    value_columns = [
        (3, "kotak_aum"),
        (4, "cams_aum"),
        (6, "kotak_gross_sales"),
        (7, "cams_gross_sales"),
        (9, "kotak_net_sales"),
        (10, "cams_net_sales"),
        (12, "kotak_sip_count"),
        (13, "cams_sip_count"),
        (15, "kotak_sip_book"),
        (16, "cams_sip_book"),
    ]
    ms_columns = [(5, 3, 4), (8, 6, 7), (11, 9, 10), (14, 12, 13), (17, 15, 16)]
    amount_columns = {3, 4, 6, 7, 9, 10, 15, 16}
    count_columns = {12, 13}
    for row_number, record in enumerate(rows, start=first_row):
        worksheet.cell(row_number, 1, record["asset_class"])
        worksheet.cell(row_number, 2, record["sch_group"])
        for column, key in value_columns:
            cell = worksheet.cell(row_number, column, record[key])
            if column in amount_columns:
                cell.number_format = AMOUNT_FMT
            elif column in count_columns:
                cell.number_format = COUNT_FMT
        for ms_column, kotak_column, cams_column in ms_columns:
            worksheet.cell(
                row_number,
                ms_column,
                f"=IFERROR(({get_column_letter(kotak_column)}{row_number}/{get_column_letter(cams_column)}{row_number}),0)",
            )
            worksheet.cell(row_number, ms_column).number_format = SHARE_FMT

    worksheet.cell(total_row, 1, "Grand Total")
    for column, _ in value_columns:
        letter = get_column_letter(column)
        cell = worksheet.cell(
            total_row, column, f"=SUBTOTAL(9,{letter}{first_row}:{letter}{expected_last})"
        )
        if column in amount_columns:
            cell.number_format = AMOUNT_FMT
        elif column in count_columns:
            cell.number_format = COUNT_FMT
    for ms_column, kotak_column, cams_column in ms_columns:
        worksheet.cell(
            total_row,
            ms_column,
            f"=IFERROR(({get_column_letter(kotak_column)}{total_row}/{get_column_letter(cams_column)}{total_row}),0)",
        )
        worksheet.cell(total_row, ms_column).number_format = SHARE_FMT

    for column in amount_columns:
        worksheet.column_dimensions[get_column_letter(column)].width = 16
    for column in count_columns:
        worksheet.column_dimensions[get_column_letter(column)].width = 13
    for ms_column, _, _ in ms_columns:
        worksheet.column_dimensions[get_column_letter(ms_column)].width = 9

    worksheet.auto_filter.ref = f"A2:Q{expected_last}"


def _write_brokerwise_sheet(worksheet, frame: pd.DataFrame) -> None:
    data_style = [copy(worksheet.cell(3, column)._style) for column in range(1, 22)]
    blank_style = [copy(worksheet.cell(48, column)._style) for column in range(1, 22)]
    total_style = [copy(worksheet.cell(49, column)._style) for column in range(1, 22)]
    if "A49:F49" in [str(value) for value in worksheet.merged_cells.ranges]:
        worksheet.unmerge_cells("A49:F49")
    worksheet.delete_rows(3, worksheet.max_row - 2)

    columns = [
        "category",
        "sub_category",
        "arn_code",
        "broker_name",
        "sch_group",
        "asset_class",
        "kotak_aum",
        "cams_aum",
        "ms_aum",
        "kotak_gross_sales",
        "cams_gross_sales",
        "ms_gross_sales",
        "kotak_net_sales",
        "cams_net_sales",
        "ms_net_sales",
        "kotak_sip_count",
        "cams_sip_count",
        "ms_sip_count",
        "kotak_sip_book",
        "cams_sip_book",
        "ms_sip_book",
    ]
    ms_positions = {9: (7, 8), 12: (10, 11), 15: (13, 14), 18: (16, 17), 21: (19, 20)}
    amount_columns = {7, 8, 10, 11, 13, 14, 19, 20}
    count_columns = {16, 17}
    text_columns = {1, 2, 3, 4, 5, 6}
    for offset, record in enumerate(frame.to_dict(orient="records"), start=3):
        for column_number, key in enumerate(columns, start=1):
            cell = worksheet.cell(offset, column_number)
            cell._style = copy(data_style[column_number - 1])
            if column_number in ms_positions:
                numerator, denominator = ms_positions[column_number]
                cell.value = (
                    f"=IFERROR(({get_column_letter(numerator)}{offset}/"
                    f"{get_column_letter(denominator)}{offset}),0)"
                )
                cell.number_format = "0.00%"
            else:
                value = record[key]
                cell.value = _safe_cell_text(value) if column_number in text_columns else value
                if column_number in amount_columns:
                    cell.number_format = AMOUNT_FMT
                elif column_number in count_columns:
                    cell.number_format = COUNT_FMT

    blank_row = 3 + len(frame)
    total_row = blank_row + 1
    for column in range(1, 22):
        worksheet.cell(blank_row, column)._style = copy(blank_style[column - 1])
        worksheet.cell(total_row, column)._style = copy(total_style[column - 1])
    worksheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
    worksheet.cell(total_row, 1, "Grand Total")
    worksheet.cell(total_row, 1).alignment = Alignment(horizontal="center")
    last_data_row = blank_row - 1
    for column in (7, 8, 10, 11, 13, 14, 16, 17, 19, 20):
        letter = get_column_letter(column)
        cell = worksheet.cell(
            total_row,
            column,
            f"=SUBTOTAL(9,{letter}3:{letter}{last_data_row})" if len(frame) else 0,
        )
        if column in amount_columns:
            cell.number_format = AMOUNT_FMT
        elif column in count_columns:
            cell.number_format = COUNT_FMT
    for ms_column, (numerator, denominator) in ms_positions.items():
        worksheet.cell(
            total_row,
            ms_column,
            f"=IFERROR(({get_column_letter(numerator)}{total_row}/"
            f"{get_column_letter(denominator)}{total_row}),0)",
        )
        worksheet.cell(total_row, ms_column).number_format = "0.00%"

    for column in amount_columns:
        worksheet.column_dimensions[get_column_letter(column)].width = 16
    for column in count_columns:
        worksheet.column_dimensions[get_column_letter(column)].width = 13
    for ms_column in ms_positions:
        worksheet.column_dimensions[get_column_letter(ms_column)].width = 9
    worksheet.column_dimensions[get_column_letter(4)].width = 30

    worksheet.auto_filter.ref = f"A2:U{last_data_row}" if len(frame) else "A2:U2"


def _configure_sip_pivot(workbook, data_row_count: int) -> None:
    """Keep the template PivotTable and point its cache at this week's data."""
    worksheet = _sheet_by_trimmed_name(workbook, "SIP Pivot")
    if len(worksheet._pivots) != 1:
        raise ValueError(
            "Template SIP Pivot must contain exactly one native Excel PivotTable"
        )
    pivot = worksheet._pivots[0]
    cache = pivot.cache
    source = cache.cacheSource.worksheetSource
    if source is None:
        raise ValueError("Template SIP Pivot cache has no worksheet source")
    source.sheet = _sheet_by_trimmed_name(workbook, "Brokerwise Data").title
    source.ref = f"A2:U{max(2, data_row_count + 2)}"
    # Keep this off during the headless formula-cache pass. It is enabled in the
    # final package XML afterwards, otherwise Excel begins an asynchronous pivot
    # refresh immediately and rejects calculation/save automation calls.
    cache.refreshOnLoad = False
    cache.enableRefresh = True
    cache.recordCount = data_row_count


def _recalculate_with_excel(path: Path) -> None:
    """Use installed Excel as the headless calculation engine.

    openpyxl intentionally does not evaluate formulas. Excel automation is the
    local equivalent of a LibreOffice headless recalc and preserves native Excel
    PivotTables while writing formula results into the workbook cache.
    """
    powershell = shutil.which("powershell") or shutil.which("pwsh")
    if not powershell:  # pragma: no cover - depends on deployment host
        raise RuntimeError(
            "Workbook recalculation requires Microsoft Excel automation or a "
            "LibreOffice headless runtime; no recalculation engine is available"
        )
    script = r"""
$ErrorActionPreference = 'Stop'
Add-Type @'
using System;
using System.Runtime.InteropServices;
public static class ExcelProcessId {
  [DllImport("user32.dll")]
  public static extern uint GetWindowThreadProcessId(IntPtr hWnd, out uint processId);
}
'@
$excel = $null
$book = $null
$excelPid = 0
$failure = $null
try {
  $excel = New-Object -ComObject Excel.Application
  [void][ExcelProcessId]::GetWindowThreadProcessId([IntPtr]$excel.Hwnd, [ref]$excelPid)
  $excel.Visible = $false
  $excel.DisplayAlerts = $false
  $excel.AskToUpdateLinks = $false
  $book = $excel.Workbooks.Open($env:MIS_RECALC_PATH, 0, $false)
  $excel.CalculateFullRebuild()
  $book.Save()
  Start-Sleep -Seconds 1
} catch {
  $failure = $_.Exception.Message
} finally {
  if ($null -ne $book) {
    try { $book.Close($false) } catch { }
    try { [void][Runtime.InteropServices.Marshal]::FinalReleaseComObject($book) } catch { }
  }
  if ($null -ne $excel) {
    try { $excel.Quit() } catch { }
    try { [void][Runtime.InteropServices.Marshal]::FinalReleaseComObject($excel) } catch { }
  }
  [GC]::Collect()
  [GC]::WaitForPendingFinalizers()
  if ($excelPid -gt 0) {
    Start-Sleep -Milliseconds 250
    Get-Process -Id $excelPid -ErrorAction SilentlyContinue |
      Stop-Process -Force -ErrorAction SilentlyContinue
  }
}
if ($null -ne $failure) { [Console]::Error.WriteLine($failure); exit 1 }
exit 0
"""
    environment = os.environ.copy()
    environment["MIS_RECALC_PATH"] = str(path.resolve())
    result = subprocess.run(
        [powershell, "-NoProfile", "-NonInteractive", "-Command", script],
        capture_output=True,
        text=True,
        timeout=120,
        env=environment,
        check=False,
    )
    if result.returncode:
        details = (result.stderr or result.stdout).strip()
        raise RuntimeError(f"Excel failed to recalculate generated workbook {path}: {details}")


def _force_recalculation_properties(path: Path) -> None:
    """Restore explicit calcPr and pivot-refresh flags after cache calculation.

    Excel omits ``calcMode=auto`` because auto is its default. Updating only the
    package XML after calculation retains the cached values that openpyxl would
    otherwise discard on a second save.
    """
    with tempfile.NamedTemporaryFile(
        prefix=path.stem + "_", suffix=".xlsx", dir=path.parent, delete=False
    ) as temporary:
        temporary_path = Path(temporary.name)
    try:
        with ZipFile(path, "r") as source, ZipFile(
            temporary_path, "w", compression=ZIP_DEFLATED
        ) as destination:
            for item in source.infolist():
                data = source.read(item.filename)
                if item.filename == "xl/workbook.xml":
                    xml = data.decode("utf-8")
                    match = re.search(r"<calcPr\b[^>]*/?>", xml)
                    if match is None:
                        tag = '<calcPr calcMode="auto" fullCalcOnLoad="1" forceFullCalc="1"/>'
                        xml = xml.replace("</workbook>", tag + "</workbook>")
                    else:
                        tag = match.group(0)
                        for attribute, value in (
                            ("calcMode", "auto"),
                            ("fullCalcOnLoad", "1"),
                            ("forceFullCalc", "1"),
                        ):
                            pattern = rf'\s{attribute}="[^"]*"'
                            replacement = f' {attribute}="{value}"'
                            if re.search(pattern, tag):
                                tag = re.sub(pattern, replacement, tag)
                            else:
                                if tag.endswith("/>"):
                                    tag = tag[:-2] + replacement + "/>"
                                else:
                                    tag = tag[:-1] + replacement + ">"
                        xml = xml[: match.start()] + tag + xml[match.end() :]
                    data = xml.encode("utf-8")
                elif item.filename.startswith(
                    "xl/pivotCache/pivotCacheDefinition"
                ) and item.filename.endswith(".xml"):
                    xml = data.decode("utf-8")
                    match = re.search(r"<pivotCacheDefinition\b[^>]*>", xml)
                    if match is None:
                        raise ValueError(f"Invalid pivot cache definition: {item.filename}")
                    tag = match.group(0)
                    for attribute, value in (
                        ("refreshOnLoad", "1"),
                        ("enableRefresh", "1"),
                    ):
                        pattern = rf'\s{attribute}="[^"]*"'
                        replacement = f' {attribute}="{value}"'
                        if re.search(pattern, tag):
                            tag = re.sub(pattern, replacement, tag)
                        else:
                            tag = tag[:-1] + replacement + ">"
                    xml = xml[: match.start()] + tag + xml[match.end() :]
                    data = xml.encode("utf-8")
                destination.writestr(item, data)
        temporary_path.replace(path)
    finally:
        temporary_path.unlink(missing_ok=True)


def _assert_recalculated_workbook(
    path: Path, expected_kotak_aum: float, total_row: int
) -> None:
    formula_workbook = load_workbook(path, read_only=False, data_only=False)
    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        errors: list[str] = []
        missing_caches: list[str] = []
        for worksheet in workbook.worksheets:
            formula_sheet = formula_workbook[worksheet.title]
            for row in worksheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.upper() in FORMULA_ERRORS:
                        errors.append(f"{worksheet.title}!{cell.coordinate}={value}")
                    formula = formula_sheet[cell.coordinate].value
                    if isinstance(formula, str) and formula.startswith("=") and value is None:
                        missing_caches.append(f"{worksheet.title}!{cell.coordinate}")
        if errors:
            raise ValueError("Generated workbook contains formula errors: " + ", ".join(errors))
        if missing_caches:
            raise ValueError(
                "Generated workbook contains formulas without cached values: "
                + ", ".join(missing_caches)
            )

        pivot_sheet = _sheet_by_trimmed_name(formula_workbook, "SIP Pivot")
        if len(pivot_sheet._pivots) != 1 or not pivot_sheet._pivots[0].cache.refreshOnLoad:
            raise ValueError("Generated workbook lost its refreshable native SIP PivotTable")

        brokerwise = _sheet_by_trimmed_name(workbook, "Brokerwise Data")
        cached_total = brokerwise.cell(total_row, 7).value
        if not isinstance(cached_total, (int, float)) or not math.isclose(
            float(cached_total), float(expected_kotak_aum), rel_tol=1e-9, abs_tol=0.01
        ):
            raise ValueError(
                "Generated workbook was not recalculated: Brokerwise Grand Total AUM "
                f"cache is {cached_total!r}, expected {expected_kotak_aum:.2f}"
            )
    finally:
        workbook.close()
        formula_workbook.close()


def generate_weekly_summary(
    frame: pd.DataFrame,
    scheme_master: list[dict],
    template_path: Path,
    output_path: Path,
) -> None:
    workbook = load_workbook(template_path)
    overall = build_summary_rows(frame, scheme_master, fintech=False)
    fintech = build_summary_rows(frame, scheme_master, fintech=True)
    _write_summary_sheet(_sheet_by_trimmed_name(workbook, EXPECTED_SHEETS[0]), overall)
    _write_summary_sheet(_sheet_by_trimmed_name(workbook, EXPECTED_SHEETS[1]), fintech)
    _write_brokerwise_sheet(_sheet_by_trimmed_name(workbook, EXPECTED_SHEETS[3]), frame)
    _configure_sip_pivot(workbook, len(frame))
    # Open every sheet at a clean, header-visible view and land on the first sheet.
    apply_safe_workbook_views(workbook)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    validate_generated_workbook(output_path)
    _recalculate_with_excel(output_path)
    _force_recalculation_properties(output_path)
    _assert_recalculated_workbook(
        output_path,
        expected_kotak_aum=float(frame["kotak_aum"].sum()) if not frame.empty else 0.0,
        total_row=4 + len(frame),
    )


def validate_generated_workbook(path: Path) -> None:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        names = workbook.sheetnames
        if names != EXPECTED_SHEETS:
            raise ValueError(f"Generated sheet order is invalid: {names}")
        bank = _sheet_by_trimmed_name(workbook, EXPECTED_SHEETS[0])
        fintech = _sheet_by_trimmed_name(workbook, EXPECTED_SHEETS[1])
        if bank.max_row != 49 or fintech.max_row != 46:
            raise ValueError("Generated summary row counts do not match the template contract")
        if not str(bank["E3"].value).startswith("=IFERROR"):
            raise ValueError("Generated market-share formulas are missing")
        pivot_sheet = _sheet_by_trimmed_name(workbook, "SIP Pivot")
        if len(pivot_sheet._pivots) != 1:
            raise ValueError("Generated workbook lost the native SIP PivotTable")
    finally:
        workbook.close()
