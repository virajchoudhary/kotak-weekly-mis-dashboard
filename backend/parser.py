from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from .constants import (
    MARKET_SHARE_COLUMNS,
    NUMERIC_COLUMNS,
    RAW_SHEET_NAME,
    TEXT_COLUMNS,
)
from .mapping import apply_mapping_rules
from .validators import MISValidationError, validate_headers


@dataclass
class ParseResult:
    frame: pd.DataFrame
    warnings: list[str]
    source_sheet: str
    source_headers: list[str]


def _read_xlsx_headers(path: Path) -> tuple[str, list[object]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise MISValidationError(f"The Excel file could not be read: {exc}") from exc
    try:
        candidates = [RAW_SHEET_NAME] + [name for name in workbook.sheetnames if name != RAW_SHEET_NAME]
        last_error: MISValidationError | None = None
        for sheet_name in candidates:
            if sheet_name not in workbook.sheetnames:
                continue
            headers = [cell.value for cell in next(workbook[sheet_name].iter_rows(min_row=1, max_row=1))]
            try:
                validate_headers(headers)
                return sheet_name, headers
            except MISValidationError as exc:
                last_error = exc
        if last_error:
            raise last_error
        raise MISValidationError("No supported worksheet with a header row was found.")
    finally:
        workbook.close()


def _read_source(path: Path) -> tuple[pd.DataFrame, str, list[object]]:
    suffix = path.suffix.lower()
    try:
        if suffix == ".csv":
            frame = pd.read_csv(path, dtype=object, keep_default_na=False)
            return frame, "CSV", list(frame.columns)
        if suffix == ".xlsx":
            sheet_name, headers = _read_xlsx_headers(path)
            frame = pd.read_excel(path, sheet_name=sheet_name, dtype=object)
            return frame, sheet_name, headers
        frame = pd.read_excel(path, sheet_name=0, dtype=object)
        return frame, "Sheet1", list(frame.columns)
    except MISValidationError:
        raise
    except ImportError as exc:
        raise MISValidationError(
            "Legacy .xls support requires the xlrd package to be installed."
        ) from exc
    except Exception as exc:
        raise MISValidationError(f"The uploaded table could not be parsed: {exc}") from exc


def _parse_number(value: object) -> tuple[float, bool]:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return 0.0, True
    if isinstance(value, (int, float, np.number)):
        return float(value), True
    text = str(value).strip()
    if text.upper() in {"", "-", "–", "—", "NA", "N/A", "NULL", "NONE"}:
        return 0.0, True
    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1]
    cleaned = text.replace(",", "").replace("₹", "").strip()
    try:
        number = float(cleaned)
        return (-number if negative else number), True
    except ValueError:
        return 0.0, False


def parse_weekly_mis(
    path: Path, scheme_master: list[dict], mapping_rules: list[dict]
) -> ParseResult:
    source, sheet_name, raw_headers = _read_source(path)
    rename_map = validate_headers(raw_headers)
    source = source.rename(columns=rename_map)[list(rename_map.values())]
    source = source.dropna(how="all")
    if source.empty:
        raise MISValidationError("The upload contains no data rows.")

    errors: list[dict] = []
    for column in NUMERIC_COLUMNS:
        parsed: list[float] = []
        for row_number, value in enumerate(source[column].tolist(), start=2):
            number, valid = _parse_number(value)
            parsed.append(number)
            if not valid and len(errors) < 100:
                errors.append(
                    {
                        "field": column,
                        "row": row_number,
                        "value": str(value),
                        "message": "Expected a numeric value",
                    }
                )
        source[column] = parsed
    if errors:
        raise MISValidationError("Invalid numeric data was found.", errors)

    for column in TEXT_COLUMNS:
        source[column] = source[column].fillna("").astype(str).str.strip()
    source["sub_category"] = source["sub_category"].replace("", "NULL")

    required_text = ["category", "arn_code", "broker_name", "sch_group", "asset_class"]
    blank_errors: list[dict] = []
    for column in required_text:
        for index in source.index[source[column].eq("")].tolist()[:20]:
            blank_errors.append(
                {
                    "field": column,
                    "row": int(index) + 2,
                    "message": "Required value is blank",
                }
            )
    if blank_errors:
        raise MISValidationError("Required row values are missing.", blank_errors)

    source = apply_mapping_rules(source, mapping_rules, scheme_master)
    for ms_column, (kotak_column, cams_column) in MARKET_SHARE_COLUMNS.items():
        denominator = source[cams_column].astype(float)
        source[ms_column] = np.where(
            denominator.ne(0), source[kotak_column].astype(float) / denominator, 0.0
        )

    ordered = TEXT_COLUMNS + NUMERIC_COLUMNS + list(MARKET_SHARE_COLUMNS)
    source = source[ordered].reset_index(drop=True)
    return ParseResult(
        frame=source,
        warnings=[],
        source_sheet=sheet_name,
        source_headers=[str(value) for value in raw_headers],
    )

