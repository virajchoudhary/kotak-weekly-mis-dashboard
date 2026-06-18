from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from backend.db import Database, utc_now
from backend.excel_engine import EXPECTED_SHEETS, generate_weekly_summary
from backend.service import UploadNotFoundError, WeeklyMISService

XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

CSV_HEADER = (
    "MAIN ARN CODE,BROKER NAME,Category,Sub Category,Scheme Group,Scheme Type,"
    "K (AUM),I (AUM),K (GS),I (GS),K (NS),I (NS),"
    "K SIP Count,I Sip Count,K SIP BOOK,I SIP BOOK"
)


def _post(client, filename, content, media, label):
    return client.post(
        "/api/uploads/weekly-mis",
        files={"file": (filename, content, media)},
        data={"week_label": label},
    )


def test_invalid_extension_is_rejected(client) -> None:
    response = _post(client, "notes.txt", b"hello world", "text/plain", "ext-1")
    assert response.status_code == 422
    assert "Unsupported file type" in response.json()["detail"]["message"]


def test_empty_file_is_rejected(client) -> None:
    response = _post(client, "empty.csv", b"", "text/csv", "empty-1")
    assert response.status_code == 422
    assert "empty" in response.json()["detail"]["message"].lower()


def test_xlsx_signature_mismatch_is_rejected(client) -> None:
    # A text file renamed to .xlsx must be rejected before it reaches the parser.
    response = _post(client, "fake.xlsx", b"definitely not a zip", XLSX_MEDIA, "sig-1")
    assert response.status_code == 422
    assert "xlsx" in response.json()["detail"]["message"].lower()


def test_valid_csv_upload_reconciles(client) -> None:
    row = "ARN-1,Test Broker,BANKS,-,Equity,Large Cap Fund,100,200,10,20,5,10,3,6,1000,2000"
    content = (CSV_HEADER + "\n" + row + "\n").encode("utf-8")
    response = _post(client, "weekly.csv", content, "text/csv", "csv-1")
    assert response.status_code == 201, response.text
    dashboard = response.json()["dashboard"]
    assert dashboard["brokerwise_total"] == 1
    assert dashboard["reconciled"] is True


def test_formula_injection_is_neutralised_in_generated_excel(settings, tmp_path: Path) -> None:
    database = Database(settings.database_path)
    database.initialize()
    master = database.fetch_scheme_master()
    frame = pd.DataFrame(
        [
            {
                "category": "BANKS",
                "sub_category": "-",
                "arn_code": "ARN-1",
                "broker_name": "=cmd|' /c calc'!A1",
                "sch_group": "Equity",
                "asset_class": "Large Cap Fund",
                "kotak_aum": 100.0,
                "cams_aum": 200.0,
                "kotak_gross_sales": 0.0,
                "cams_gross_sales": 0.0,
                "kotak_net_sales": 0.0,
                "cams_net_sales": 0.0,
                "kotak_sip_count": 0.0,
                "cams_sip_count": 0.0,
                "kotak_sip_book": 0.0,
                "cams_sip_book": 0.0,
            }
        ]
    )
    output = tmp_path / "injection.xlsx"
    generate_weekly_summary(frame, master, settings.template_path, output)
    workbook = load_workbook(output)
    cell = workbook["Brokerwise Data"]["D3"]  # broker_name, first data row
    assert cell.data_type != "f", "user text must never become a live formula"
    assert str(cell.value).startswith("'="), "leading formula trigger must be escaped"
    workbook.close()


def test_download_rejects_path_outside_generated_dir(settings, tmp_path: Path) -> None:
    settings.ensure_directories()
    database = Database(settings.database_path)
    database.initialize()
    service = WeeklyMISService(settings, database)

    evil = tmp_path / "secret.xlsx"
    evil.write_bytes(b"PK\x03\x04 pretend workbook")
    now = utc_now()
    with database.transaction() as conn:
        conn.execute(
            """
            INSERT INTO uploads (
                week_label, week_start_date, week_end_date, upload_date,
                original_filename, file_hash, raw_file_path, generated_file_path,
                row_count, status, validation_summary_json, created_at, updated_at
            ) VALUES ('evil-wk', NULL, NULL, ?, 'x.xlsx', 'hash-evil', '', ?, 0,
                      'finalized', '{}', ?, ?)
            """,
            (now, str(evil.resolve()), now, now),
        )
        upload_id = conn.execute(
            "SELECT id FROM uploads WHERE week_label='evil-wk'"
        ).fetchone()[0]

    with pytest.raises(UploadNotFoundError):
        service.download_path(upload_id)


def test_download_nonexistent_upload_is_404(client) -> None:
    assert client.get("/api/download/999999").status_code == 404


def test_generated_workbook_still_has_exactly_four_sheets(settings, tmp_path: Path) -> None:
    database = Database(settings.database_path)
    database.initialize()
    master = database.fetch_scheme_master()
    from backend.parser import parse_weekly_mis
    from .conftest import SAMPLE_PATH

    frame = parse_weekly_mis(SAMPLE_PATH, master, database.fetch_mapping_rules()).frame
    output = tmp_path / "sheets.xlsx"
    generate_weekly_summary(frame, master, settings.template_path, output)
    workbook = load_workbook(output)
    assert workbook.sheetnames == EXPECTED_SHEETS
    workbook.close()
