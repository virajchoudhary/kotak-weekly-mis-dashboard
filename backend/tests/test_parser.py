from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend.db import Database
from backend.parser import parse_weekly_mis
from backend.validators import MISValidationError
from .conftest import SAMPLE_PATH


def test_parses_provided_weekly_mis(settings) -> None:
    database = Database(settings.database_path)
    database.initialize()
    result = parse_weekly_mis(
        SAMPLE_PATH, database.fetch_scheme_master(), database.fetch_mapping_rules()
    )
    assert result.source_sheet == "Brokerwise Data (67)"
    assert len(result.frame) == 45
    assert result.frame.iloc[0]["asset_class"] == "Banking and PSU Fund"
    assert result.frame["ms_aum"].eq(0).all()


def test_column_mapping_and_ms_calculation(settings, tmp_path: Path) -> None:
    source = tmp_path / "calculation.xlsx"
    workbook = load_workbook(SAMPLE_PATH)
    sheet = workbook["Brokerwise Data (67)"]
    sheet["G2"] = 125
    sheet["H2"] = 500
    sheet["J2"] = 10
    sheet["K2"] = 0
    workbook.save(source)

    database = Database(settings.database_path)
    database.initialize()
    result = parse_weekly_mis(
        source, database.fetch_scheme_master(), database.fetch_mapping_rules()
    )
    first = result.frame.iloc[0]
    assert first["arn_code"] == "ARN-0000"
    assert first["kotak_aum"] == 125
    assert first["cams_aum"] == 500
    assert first["ms_aum"] == pytest.approx(0.25)
    assert first["ms_gross_sales"] == 0


def test_missing_required_column_is_rejected(settings, tmp_path: Path) -> None:
    source = tmp_path / "invalid.xlsx"
    workbook = load_workbook(SAMPLE_PATH)
    sheet = workbook["Brokerwise Data (67)"]
    sheet.delete_cols(7)
    workbook.save(source)
    database = Database(settings.database_path)
    database.initialize()
    with pytest.raises(MISValidationError, match="Required columns"):
        parse_weekly_mis(
            source, database.fetch_scheme_master(), database.fetch_mapping_rules()
        )


def test_invalid_numeric_value_is_rejected(settings, tmp_path: Path) -> None:
    source = tmp_path / "invalid-number.xlsx"
    workbook = load_workbook(SAMPLE_PATH)
    workbook["Brokerwise Data (67)"]["G2"] = "not-a-number"
    workbook.save(source)
    database = Database(settings.database_path)
    database.initialize()
    with pytest.raises(MISValidationError, match="Invalid numeric"):
        parse_weekly_mis(
            source, database.fetch_scheme_master(), database.fetch_mapping_rules()
        )

