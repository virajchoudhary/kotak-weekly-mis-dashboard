from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend.db import Database
from backend.excel_engine import (
    EXPECTED_SHEETS,
    _configure_sip_pivot,
    _force_recalculation_properties,
    build_summary_rows,
    generate_weekly_summary,
)
from backend.parser import parse_weekly_mis
from .conftest import SAMPLE_PATH


def _parsed(settings):
    database = Database(settings.database_path)
    database.initialize()
    master = database.fetch_scheme_master()
    parsed = parse_weekly_mis(SAMPLE_PATH, master, database.fetch_mapping_rules())
    return parsed.frame, master


def test_summary_generation_has_45_and_42_rows(settings) -> None:
    frame, master = _parsed(settings)
    banks = build_summary_rows(frame, master, fintech=False)
    fintech = build_summary_rows(frame, master, fintech=True)
    assert len(banks) == 45
    assert len(fintech) == 42
    excluded = {"Capital Protection Oriented Schemes", "INCOME/DEBT (INTERVAL)", "OTHER DEBT (C)"}
    assert excluded.isdisjoint({row["asset_class"] for row in fintech})
    assert sum(row["kotak_aum"] for row in banks) == sum(frame["kotak_aum"])


def test_generated_workbook_opens_and_preserves_contract(settings, tmp_path: Path) -> None:
    frame, master = _parsed(settings)
    output = tmp_path / "generated.xlsx"
    generate_weekly_summary(frame, master, settings.template_path, output)
    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == EXPECTED_SHEETS
    assert workbook[EXPECTED_SHEETS[0]].max_row == 49
    assert workbook[EXPECTED_SHEETS[1]].max_row == 46
    assert workbook[EXPECTED_SHEETS[0]]["E3"].value == "=IFERROR((C3/D3),0)"
    assert workbook[EXPECTED_SHEETS[0]]["A1"].value == "Summary - Banks, ND & RIA"
    assert workbook[EXPECTED_SHEETS[3]]["I3"].value == "=IFERROR((G3/H3),0)"
    assert workbook.calculation.calcMode == "auto"
    assert workbook.calculation.fullCalcOnLoad is True
    pivot_sheet = workbook[EXPECTED_SHEETS[2]]
    assert len(pivot_sheet._pivots) == 1
    pivot_cache = pivot_sheet._pivots[0].cache
    assert pivot_cache.refreshOnLoad is True
    assert pivot_cache.cacheSource.worksheetSource.sheet == EXPECTED_SHEETS[3]
    assert pivot_cache.cacheSource.worksheetSource.ref == f"A2:U{len(frame) + 2}"
    workbook.close()

    cached = load_workbook(output, data_only=True)
    brokerwise_total_row = len(frame) + 4
    assert cached[EXPECTED_SHEETS[3]].cell(brokerwise_total_row, 7).value == pytest.approx(
        frame["kotak_aum"].sum()
    )
    assert cached[EXPECTED_SHEETS[0]]["E3"].value is not None
    formula_errors = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?"}
    assert not [
        (ws.title, cell.coordinate, cell.value)
        for ws in cached.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.upper() in formula_errors
    ]
    cached.close()


def test_sip_pivot_cache_source_tracks_variable_week_size(settings, tmp_path: Path) -> None:
    workbook = load_workbook(settings.template_path)
    _configure_sip_pivot(workbook, data_row_count=7)
    output = tmp_path / "dynamic-pivot.xlsx"
    workbook.save(output)
    workbook.close()
    _force_recalculation_properties(output)

    generated = load_workbook(output, data_only=False)
    pivot = generated[EXPECTED_SHEETS[2]]._pivots[0]
    assert pivot.cache.cacheSource.worksheetSource.ref == "A2:U9"
    assert pivot.cache.refreshOnLoad is True
    generated.close()


def test_generated_workbook_opens_at_clean_top_left_view(settings, tmp_path: Path) -> None:
    """Every generated sheet must open at the top-left with visible headers, the
    first sheet active, and no frozen panes — without disturbing data."""
    frame, master = _parsed(settings)
    output = tmp_path / "generated.xlsx"
    generate_weekly_summary(frame, master, settings.template_path, output)
    workbook = load_workbook(output, data_only=False)

    # Exactly the four expected sheets, in order, and the first one opens first.
    assert workbook.sheetnames == EXPECTED_SHEETS
    assert workbook.active.title == EXPECTED_SHEETS[0]

    for index, title in enumerate(EXPECTED_SHEETS):
        ws = workbook[title]
        view = ws.sheet_view
        assert ws.freeze_panes is None, f"{title} freeze {ws.freeze_panes}"
        assert view.topLeftCell in (None, "A1"), f"{title} opens scrolled at {view.topLeftCell}"
        assert view.selection and view.selection[0].activeCell == "A1", f"{title} active cell"
        # Only the first sheet is tab-selected.
        assert bool(view.tabSelected) == (index == 0), f"{title} tabSelected={view.tabSelected}"

    # Data/formulas/totals remain intact after the view normalization.
    assert workbook[EXPECTED_SHEETS[0]]["E3"].value == "=IFERROR((C3/D3),0)"
    assert str(workbook[EXPECTED_SHEETS[0]]["C49"].value).startswith("=SUBTOTAL")
    workbook.close()
